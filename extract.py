#!/usr/bin/python
#
TABLES = 'Tables'
from secure import USER, PASS
import os
import simplejson as json
from datetime import datetime

#Sample JSON Table Structure
# {
#     "id": "Table",
#     "server": "SERVER",
#     "database": "DB",
#     "query": "SELECT * FROM dbo.XYZ",
#     "fields": ["COLUMN1","COLUMN2"],
#     "mappings": ["Id","Dog"]
# }


class Table(object):

    def __init__(self, filename):
        self.filename = filename
        self.json = None
        with open(os.path.join(os.getcwd(), TABLES, filename)) as f:
            self.json = json.load(f)
        self.server = self.json["server"]
        self.database = self.json["database"]
        self.query = self.json["query"]
        self.fields = self.json["fields"]
        self.mappings = self.json["mappings"]
        self.id = self.json["id"]
        self.query = self.query.replace("*", ",".join(self.fields))

    def extract(self):
        """ Extracts, duh.
        """
        from sqlalchemy import create_engine
        engine = create_engine("mssql+pyodbc://%s:%s@%s/%s" % (
            USER, PASS, self.server, self.database))
        #
        result = engine.execute(self.query)
        from openpyxl import Workbook
        from openpyxl.cell import get_column_letter
        wb = Workbook()
        ws = wb.active
        for idx in xrange(1, 3):
            for col_idx in xrange(1, len(self.fields)+1):
                col = get_column_letter(col_idx)
                if idx == 1:
                    ws.cell('%s%s' % (col, idx)).value = self.fields[col_idx-1]
                elif idx == 2:
                    ws.cell('%s%s' % (col, idx)).value = self.mappings[col_idx-1]
        #
        for idx, row in enumerate(result):
            for col_idx in xrange(1, len(row)+1):
                col = get_column_letter(col_idx)
                ws.cell('%s%s' % (col, idx+3)).value = str(row[col_idx-1]).decode('utf-8', 'ignore')
        ws.title = "Extract"
        ws = wb.create_sheet()

        wb.save(filename="Extracts\%s-%s.xls" %
                (self.id, datetime.now().strftime("%y%m%d")))

if __name__ == "__main__":
    for filename in os.listdir(os.path.join(os.getcwd(), TABLES)):
        table = Table(filename)
        table.extract()
